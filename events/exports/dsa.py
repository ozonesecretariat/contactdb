import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from events.models import Event, Registration

ROW_HEIGHTS = {
    5: 55,
}

COL_WIDTHS = {
    "A": 2,
    "B": 4,
    "C": 21,
    "D": 13,
    "E": 16,
    "F": 10,
    "G": 16,
    "H": 20,
    "I": 14,
    "J": 12,
    "K": 6,
    "L": 8,
    "M": 6,
    "N": 10,
    "O": 30,
}


class DSAReport:
    SUFFIX = "-DSA-report"

    def __init__(self, event: Event):
        self.event = event
        self.wb = Workbook()
        self.ws = self.wb.active
        self.start_row = 1
        self.start_col = 2
        self.row_index = self.start_row
        self.max_col = self.start_col
        self.max_header_row = self.start_row
        self.row_font = Font(name="Arial", size=10)
        self.row_align = Alignment(wrap_text=True)
        self.header_font = Font(bold=True, name="Arial", size=10)
        self.header_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.header_bg = PatternFill(start_color="99ccff", fill_type="solid")
        self.registrations = (
            self.event.registrations.filter(
                status=Registration.Status.REGISTERED,
                tags__name="Is funded",
                dsa__isnull=False,
            )
            .select_related("dsa")
            .prefetch_related(
                "role",
                "tags",
                "event",
                "event__venue_country",
                "organization",
                "organization__government",
                "organization__country",
                "organization__organization_type",
                "contact",
                "contact__country",
                "contact__organization",
                "contact__organization__government",
                "contact__organization__country",
                "contact__organization__organization_type",
            )
            .order_by("contact__last_name", "contact__first_name")
        )

    def export_xlsx(self):
        self.set_dimensions()
        self.write_headers(
            [
                [
                    {"label": self.event.title, "colspan": 14},
                ],
                [
                    {
                        "label": ", ".join(
                            [
                                self.event.dates,
                                self.event.venue_city,
                                self.event.venue_country.name,
                            ]
                        ),
                        "colspan": 14,
                    },
                ],
                [
                    {"label": "", "colspan": 9},
                    {"label": "DSA rate in USD", "colspan": 2},
                    {"label": self.event.dsa},
                    {"label": "", "colspan": 2},
                ],
            ],
            bg_color=False,
        )
        self.write_headers(
            [
                [
                    {"label": "N° of PAX", "rowspan": 2},
                    {"label": "Country", "rowspan": 2},
                    {"label": "Umoja Travel #", "rowspan": 2},
                    {"label": "BP #", "rowspan": 2},
                    {"label": "Participant's name", "colspan": 3},
                    {"label": "Period covered", "colspan": 2},
                    {"label": "N° of days", "rowspan": 2},
                    {"label": "100% of DSA on arrival date in USD", "rowspan": 2},
                    {"label": "Term. Exp. in USD", "rowspan": 2},
                    {"label": "Total DSA in USD", "rowspan": 2},
                    {"label": "Signature upon receipt of total", "rowspan": 2},
                ],
                [
                    {"colspan": 4},
                    {"label": "Salutation (Mr/Mrs)"},
                    {"label": "Family Name"},
                    {"label": "First Name"},
                    {"label": "Arrival date\ndd/mm/yyyy"},
                    {"label": "Departure date\ndd/mm/yyyy"},
                ],
            ],
            bg_color=True,
        )
        self.write_headers(
            [
                [
                    {
                        "label": "Sponsored members / Coding block / Event ID Number",
                        "colspan": 14,
                    },
                ]
            ],
            bg_color=False,
        )
        self.write_data()
        self.write_footer()
        self.set_border()
        self.ws.freeze_panes = f"A{self.max_header_row + 1}"
        return self.save_xlsx()

    def set_dimensions(self):
        for row, height in ROW_HEIGHTS.items():
            self.ws.row_dimensions[row].height = height
        for col, width in COL_WIDTHS.items():
            self.ws.column_dimensions[col].width = width

    def write_headers(self, headers, bg_color=True):
        for row in headers:
            col_index = self.start_col
            for cell_data in row:
                label = cell_data.get("label")
                colspan = cell_data.get("colspan", 1)
                rowspan = cell_data.get("rowspan", 1)
                if label is None:
                    # Skip this cell
                    col_index += colspan
                    continue

                cell = self.ws.cell(row=self.row_index, column=col_index, value=label)
                cell.font = self.header_font
                cell.alignment = self.header_align
                if bg_color:
                    cell.fill = self.header_bg
                self.ws.merge_cells(
                    start_row=self.row_index,
                    start_column=col_index,
                    end_row=self.row_index + (rowspan - 1),
                    end_column=col_index + (colspan - 1),
                )
                col_index += colspan
                self.max_col = max(self.max_col, col_index)
            self.max_header_row = self.row_index
            self.row_index += 1

    def write_data(self):
        for reg_index, reg in enumerate(self.registrations, start=1):
            self.write_row(
                [
                    reg_index,
                    reg.dsa_country and reg.dsa_country.name,
                    reg.dsa.umoja_travel,
                    reg.dsa.bp,
                    reg.contact.title,
                    reg.contact.last_name,
                    reg.contact.first_name,
                    {"value": reg.dsa.arrival_date, "number_format": "DD/MM/YYYY"},
                    {"value": reg.dsa.departure_date, "number_format": "DD/MM/YYYY"},
                    f"=J{self.row_index}-I{self.row_index}",
                    f"=M3*K{self.row_index}",
                    reg.event.term_exp,
                    f"=L{self.row_index}+M{self.row_index}",
                ]
            )

    def write_row(self, row):
        self.ws.row_dimensions[self.row_index].height = 30
        for col_index, cell_data in enumerate(row, start=self.start_col):
            if not isinstance(cell_data, dict):
                cell_data = {"value": cell_data}

            value = cell_data.get("value", "") or ""
            cell = self.ws.cell(row=self.row_index, column=col_index, value=value)
            cell.font = self.row_font
            cell.alignment = self.row_align
            if "number_format" in cell_data:
                cell.number_format = cell_data["number_format"]

        self.row_index += 1

    def write_footer(self):
        cell = self.ws.cell(
            row=self.row_index,
            column=14,
            value=f"=SUM(N{self.max_header_row + 1}:N{self.row_index - 1})",
        )
        cell.font = self.header_font
        self.row_index += 1

    def set_border(self):
        for row_index in range(self.start_row, self.row_index):
            for col_index in range(self.start_col, self.max_col):
                cell = self.ws.cell(row=row_index, column=col_index)

                top = "medium" if row_index == self.start_row else "thin"
                right = "medium" if col_index == self.max_col - 1 else "thin"
                bottom = "medium" if row_index == self.row_index - 1 else "thin"
                left = "medium" if col_index == self.start_col else "thin"

                cell.border = Border(
                    top=Side(color="000000", border_style=top),
                    right=Side(color="000000", border_style=right),
                    bottom=Side(color="000000", border_style=bottom),
                    left=Side(color="000000", border_style=left),
                )

    def save_xlsx(self):
        xlsx_file = io.BytesIO()
        xlsx_file.name = f"{self.event.code}{self.SUFFIX}.xlsx"
        self.wb.save(xlsx_file)
        xlsx_file.seek(0)
        return xlsx_file
