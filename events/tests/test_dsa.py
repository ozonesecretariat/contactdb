import io
import zipfile
from datetime import datetime, timedelta

from django.conf import settings
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from api.tests.factories import (
    ContactFactory,
    CountryFactory,
    DSAFactory,
    EventFactory,
    OrganizationFactory,
    RegistrationFactory,
)
from core.models import OrganizationType
from core.templatetags.file_base64 import file_to_base64
from events.models import Registration, RegistrationTag


class TestDSAReport(TestCase):
    fixtures = [
        "initial/organizationtype",
        "initial/registrationtag",
        "initial/role",
        "test/user",
    ]

    def setUp(self):
        self.maxDiff = None
        self.event = EventFactory()

        self.is_funded = RegistrationTag.objects.get(name="Is funded")

        self.country0 = CountryFactory(name="Country 0")
        self.country1 = CountryFactory(name="Country 1")
        self.country2 = CountryFactory(name="Country 2")

        self.gov = OrganizationType.objects.get(acronym="GOV")
        self.ass = OrganizationType.objects.get(acronym="ASSMT-PANEL")

        self.org1 = OrganizationFactory(
            government=self.country1, organization_type=self.gov
        )
        self.org2 = OrganizationFactory(
            country=self.country2, organization_type=self.ass
        )

        self.contact1 = ContactFactory(organization=self.org1)
        self.contact2 = ContactFactory(organization=self.org2)

        self.reg1 = RegistrationFactory(
            event=self.event,
            contact=self.contact1,
            organization=self.org1,
            status=Registration.Status.REGISTERED,
        )
        self.reg1.tags.add(self.is_funded)
        self.reg2 = RegistrationFactory(
            event=self.event,
            contact=self.contact2,
            organization=self.org2,
            status=Registration.Status.REGISTERED,
        )
        self.reg2.tags.add(self.is_funded)

        self.today = timezone.now().date()
        self.date1 = datetime.combine(self.today, datetime.min.time())
        self.date2 = datetime.combine(
            self.date1 + timedelta(days=7), datetime.min.time()
        )

        self.dsa1 = DSAFactory(
            registration=self.reg1, arrival_date=self.date1, departure_date=self.date2
        )
        self.dsa2 = DSAFactory(
            registration=self.reg2, arrival_date=self.date1, departure_date=self.date2
        )

    def check_report(self, expected=None, max_col=10):
        url = (
            reverse("registration-export-dsa")
            + f"?eventCode={self.event.code}&status=Registered,Accredited&tag=Is+funded"
        )
        self.client.login(email="admin@example.com", password="admin")
        resp = self.client.get(url)
        f = io.BytesIO(b"".join(resp.streaming_content))
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active

        result = []
        for row in ws.iter_rows(min_row=7):
            row_result = []
            for cell in row[1:max_col]:
                row_result.append(cell.value)

            result.append(row_result)

        if expected:
            self.assertEqual(result[:-1], expected)
        return result[:-1]

    def check_contacts(self, expected_contacts):
        result = self.check_report()

        self.assertEqual(len(result), len(expected_contacts))
        for row, contact in zip(result, expected_contacts, strict=False):
            self.assertEqual(row[5], contact.last_name)

        return result

    def test_dsa_report(self):
        self.check_report(
            [
                [
                    1,
                    self.country1.name,
                    str(self.dsa1.umoja_travel),
                    str(self.dsa1.bp),
                    str(self.contact1.title),
                    self.contact1.last_name,
                    self.contact1.first_name,
                    self.date1,
                    self.date2,
                ],
                [
                    2,
                    self.country2.name,
                    str(self.dsa2.umoja_travel),
                    str(self.dsa2.bp),
                    str(self.contact2.title),
                    self.contact2.last_name,
                    self.contact2.first_name,
                    self.date1,
                    self.date2,
                ],
            ]
        )

    def test_dsa_missing(self):
        self.dsa1.delete()
        self.check_contacts([self.contact1, self.contact2])

    def test_not_funded(self):
        self.reg1.tags.remove(self.is_funded)
        self.check_contacts([self.contact2])

    def test_not_registered(self):
        self.reg1.status = Registration.Status.REVOKED
        self.reg1.save()

        self.check_contacts([self.contact2])

    def test_no_country(self):
        self.org1.government = None
        self.org1.save()
        self.check_contacts([self.contact2, self.contact1])

    def test_use_org_address(self):
        self.org2.country = self.country0
        self.org2.save()
        self.contact2.is_use_organization_address = True
        self.contact2.save()

        result = self.check_contacts([self.contact2, self.contact1])
        self.assertEqual(result[0][1], self.country0.name)
        self.assertEqual(result[1][1], self.country1.name)

    def test_use_contact_address(self):
        self.contact2.is_use_organization_address = False
        self.contact2.country = self.country0
        self.contact2.save()

        result = self.check_contacts([self.contact2, self.contact1])
        self.assertEqual(result[0][1], self.country0.name)
        self.assertEqual(result[1][1], self.country1.name)

    def test_no_dates(self):
        self.dsa1.arrival_date = None
        self.dsa1.departure_date = None
        self.dsa1.save()
        self.check_contacts([self.contact1, self.contact2])


class TestDSAFiles(TestCase):
    fixtures = [
        "initial/organizationtype",
        "initial/registrationtag",
        "initial/role",
        "test/user",
    ]

    def setUp(self):
        self.maxDiff = None
        self.event = EventFactory(code="TESTEVT")

        with (settings.BASE_DIR / "fixtures/test/files/test-logo.png").open("rb") as f:
            self.data_uri = file_to_base64(f)
        self.file_obj = {"data": self.data_uri, "filename": "test-logo.png"}

        self.is_funded = RegistrationTag.objects.get(name="Is funded")

        self.reg1 = RegistrationFactory(
            event=self.event, status=Registration.Status.REGISTERED
        )
        self.reg1.tags.add(self.is_funded)
        self.dsa1 = DSAFactory(
            registration=self.reg1, passport=self.file_obj, boarding_pass=self.file_obj
        )

    def check_report(self, expected_files):
        url = (
            reverse("registration-export-dsa-files")
            + f"?eventCode={self.event.code}&status=Registered&tag=Is+funded"
        )
        self.client.login(email="admin@example.com", password="admin")
        resp = self.client.get(url)
        f = io.BytesIO(b"".join(resp.streaming_content))
        files = sorted(zipfile.ZipFile(f, mode="r").namelist())
        self.assertEqual(files, expected_files)
        return f

    def test_dsa_files_export(self):
        reg2 = RegistrationFactory(
            event=self.event, status=Registration.Status.REGISTERED
        )
        reg2.tags.add(self.is_funded)
        DSAFactory(
            registration=reg2, passport=self.file_obj, boarding_pass=self.file_obj
        )

        self.check_report(
            [
                f"{self.reg1.contact.last_name}_{self.reg1.contact.first_name}_{self.reg1.contact.id}_boarding_pass.pdf",
                f"{self.reg1.contact.last_name}_{self.reg1.contact.first_name}_{self.reg1.contact.id}_passport.pdf",
                f"{reg2.contact.last_name}_{reg2.contact.first_name}_{reg2.contact.id}_boarding_pass.pdf",
                f"{reg2.contact.last_name}_{reg2.contact.first_name}_{reg2.contact.id}_passport.pdf",
            ]
        )

    def test_export_with_missing_files(self):
        self.dsa1.boarding_pass = None
        self.dsa1.save()
        zipf = self.check_report(
            [
                f"{self.reg1.contact.last_name}_{self.reg1.contact.first_name}_{self.reg1.contact.id}_passport.pdf"
            ]
        )
        with zipfile.ZipFile(zipf) as z:
            pdf_bytes = z.read(z.namelist()[0])
            self.assertTrue(pdf_bytes.startswith(b"%PDF"))

    def test_export_with_no_files(self):
        self.dsa1.passport = None
        self.dsa1.boarding_pass = None
        self.dsa1.save()
        self.check_report([])

    def test_not_registered(self):
        self.reg1.status = Registration.Status.REVOKED
        self.reg1.save()

        self.check_report([])

    def test_not_funded(self):
        self.reg1.tags.remove(self.is_funded)
        self.reg1.save()

        self.check_report([])
