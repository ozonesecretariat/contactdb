import io
import pytest
import openpyxl

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from core.models import Record

pytestmark = [pytest.mark.django_db]


def test_import_excel(login_user_can_import, contact):
    client, user = login_user_can_import
    contact.organization.save()
    contact.first_name = "First"
    contact.last_name = "Last"
    contact.save()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    for col in range(1, 25):
        ws.cell(row=1, column=col, value="Header")
        ws.cell(row=2, column=col, value="Values")
    ws.cell(row=2, column=9, value="Other")
    ws.cell(row=2, column=10, value="Contact")

    excel_file = io.BytesIO()
    wb.save(excel_file)
    url = reverse("import-contacts")
    response = client.post(
        url,
        {
            "uploaded-excel": SimpleUploadedFile(
                name="test_file", content=excel_file.getvalue()
            )
        },
    )
    assert response.status_code == 302
    assert Record.objects.get(first_name="Other", last_name="Contact")
